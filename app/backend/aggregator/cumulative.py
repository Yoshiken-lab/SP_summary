"""
累積集計モジュール
web_app の SalesSummaryProcessor から移植・リファクタリング
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import logging

logger = logging.getLogger(__name__)


class CumulativeAggregator:
    """
    累積集計クラス
    月次集計結果を年度累積表に追記する

    使用例:
        aggregator = CumulativeAggregator(input_path, output_path, year=2024, month=12)
        result = aggregator.process()
    """

    def __init__(
        self,
        input_path: Path,
        output_dir: Path,
        year: int,
        month: int,
        fiscal_year: int,
        existing_file_path: Path = None
    ):
        """
        Args:
            input_path: 月次集計結果Excelファイルのパス
            output_dir: 出力先ディレクトリ
            year: 対象年（例: 2024）
            month: 対象月（例: 12）
            fiscal_year: 年度（例: 2024 → 2024年度）
            existing_file_path: 既存の累積ファイルパス（指定時はそのファイルに追記）
        """
        self.input_path = Path(input_path)
        self.output_dir = Path(output_dir)
        self.year = year
        self.month = month
        self.fiscal_year = fiscal_year
        self.existing_file_path = Path(existing_file_path) if existing_file_path else None

        # 月カラム名
        self.month_col_name = f"{year}年{month}月分"

        # 出力ファイル名・パス
        if self.existing_file_path and self.existing_file_path.exists():
            # 既存ファイルが指定された場合、そのファイルに上書き
            self.output_path = self.existing_file_path
            self.output_filename = self.existing_file_path.name
        else:
            # 新規作成の場合
            self.output_filename = f"SP_年度累計_{fiscal_year}.xlsx"
            self.output_path = self.output_dir / self.output_filename

        # 結果
        self.school_count = 0
        self.event_count = 0

    def process(self) -> dict:
        """
        累積集計を実行

        Returns:
            dict: 処理結果
        """
        logger.info(f"累積集計開始: {self.month_col_name}")
        logger.info(f"入力ファイル: {self.input_path}")
        logger.info(f"出力ファイル: {self.output_path}")

        # 入力ファイル読み込み
        input_xl = pd.ExcelFile(self.input_path)

        # シート存在チェック
        if '学校別' not in input_xl.sheet_names:
            raise ValueError("入力ファイルに「学校別」シートがありません")
        if 'イベント別' not in input_xl.sheet_names:
            raise ValueError("入力ファイルに「イベント別」シートがありません")

        df_school = pd.read_excel(input_xl, sheet_name='学校別')
        df_event = pd.read_excel(input_xl, sheet_name='イベント別')

        logger.info(f"学校別: {len(df_school)}件")
        logger.info(f"学校別カラム: {list(df_school.columns)}")
        logger.info(f"イベント別: {len(df_event)}件")
        logger.info(f"イベント別カラム: {list(df_event.columns)}")

        # カラム名の正規化（「売り上げ」と「売上」の両方に対応）
        df_school = self._normalize_columns(df_school)
        df_event = self._normalize_columns(df_event)

        # 出力ファイルの処理
        if self.output_path.exists() and self._is_valid_existing_file():
            logger.info("既存ファイルに追記します")
            self._update_existing_file(df_school, df_event)
        else:
            logger.info("新規ファイルを作成します")
            self._create_new_file(df_school, df_event)

        logger.info("累積集計完了")

        return {
            'status': 'success',
            'schoolCount': self.school_count,
            'eventCount': self.event_count,
            'outputPath': str(self.output_path),
            'outputFilename': self.output_filename
        }

    def _normalize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """カラム名を正規化"""
        # カラム名のマッピング（入力で見つかる可能性のある名前 → 標準名）
        column_mapping = {
            '売上': '売り上げ',
            '売上げ': '売り上げ',
            '売上額': '売り上げ',
        }

        # 存在するカラムのみリネーム
        rename_dict = {}
        for old_name, new_name in column_mapping.items():
            if old_name in df.columns and new_name not in df.columns:
                rename_dict[old_name] = new_name

        if rename_dict:
            logger.info(f"カラム名を正規化: {rename_dict}")
            df = df.rename(columns=rename_dict)

        return df

    def _is_valid_existing_file(self) -> bool:
        """既存ファイルが有効なデータを持っているかチェック"""
        try:
            xl = pd.ExcelFile(self.output_path)
            if '学校別' not in xl.sheet_names:
                return False
            df = pd.read_excel(xl, sheet_name='学校別')
            if df.empty or '担当者' not in df.columns:
                return False
            return True
        except Exception:
            return False

    def _create_new_file(self, df_school: pd.DataFrame, df_event: pd.DataFrame):
        """新規ファイルを作成"""
        # 学校別シートのデータ準備
        school_cols = ['担当者', '写真館', '学校名', '売り上げ']
        school_data = df_school[[c for c in school_cols if c in df_school.columns]].copy()
        school_data = school_data.rename(columns={'売り上げ': self.month_col_name})
        school_data['総計'] = school_data[self.month_col_name]

        # イベント別シートのデータ準備（事業所がない場合も対応）
        event_cols_with_branch = ['事業所', '学校名', 'イベント名', 'イベント開始日', '売り上げ']
        event_cols_without_branch = ['学校名', 'イベント名', 'イベント開始日', '売り上げ']

        if '事業所' in df_event.columns:
            event_cols = event_cols_with_branch
        else:
            event_cols = event_cols_without_branch

        event_data = df_event[[c for c in event_cols if c in df_event.columns]].copy()
        event_data = event_data.rename(columns={'売り上げ': self.month_col_name})
        event_data['総計'] = event_data[self.month_col_name]

        # Excelファイルに書き込み
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            school_data.to_excel(writer, sheet_name='学校別', index=False)
            event_data.to_excel(writer, sheet_name='イベント別', index=False)

        # スタイル適用
        self._apply_styles()

        self.school_count = len(school_data)
        self.event_count = len(event_data)

        logger.info(f"学校別シート: {self.school_count}行")
        logger.info(f"イベント別シート: {self.event_count}行")

    def _update_existing_file(self, df_school: pd.DataFrame, df_event: pd.DataFrame):
        """既存ファイルに月列を追加"""
        # 既存ファイル読み込み
        existing_school = pd.read_excel(self.output_path, sheet_name='学校別')
        existing_event = pd.read_excel(self.output_path, sheet_name='イベント別')

        # 学校別の更新
        updated_school = self._merge_school_data(existing_school, df_school)

        # イベント別の更新
        updated_event = self._merge_event_data(existing_event, df_event)

        # Excelファイルに書き込み
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            updated_school.to_excel(writer, sheet_name='学校別', index=False)
            updated_event.to_excel(writer, sheet_name='イベント別', index=False)

        # スタイル適用
        self._apply_styles()

        self.school_count = len(updated_school)
        self.event_count = len(updated_event)

        logger.info(f"学校別シート: {self.school_count}行")
        logger.info(f"イベント別シート: {self.event_count}行")

    def _merge_school_data(self, existing_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
        """学校別データをマージ"""
        new_data = new_df[['担当者', '写真館', '学校名', '売り上げ']].copy()
        new_data = new_data.rename(columns={'売り上げ': self.month_col_name})

        if '総計' in existing_df.columns:
            existing_df = existing_df.drop(columns=['総計'])

        if self.month_col_name in existing_df.columns:
            existing_df = existing_df.drop(columns=[self.month_col_name])

        key_cols = ['担当者', '写真館', '学校名']
        merged = pd.merge(existing_df, new_data, on=key_cols, how='outer')
        merged = self._reorder_month_columns(merged, key_cols, sheet_type='school')

        month_cols = [col for col in merged.columns if '年' in col and '月分' in col]
        merged['総計'] = merged[month_cols].fillna(0).sum(axis=1).astype(int)

        return merged

    def _merge_event_data(self, existing_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
        """イベント別データをマージ"""
        # 事業所の有無を確認
        new_has_branch = '事業所' in new_df.columns
        existing_has_branch = '事業所' in existing_df.columns

        # 新しいデータから事業所マッピングを作成（学校名 → 事業所）
        branch_mapping = {}
        if new_has_branch:
            for _, row in new_df[['学校名', '事業所']].drop_duplicates().iterrows():
                if pd.notna(row['事業所']):
                    branch_mapping[row['学校名']] = row['事業所']

        # 既存ファイルに事業所カラムがない場合、追加して埋める
        if not existing_has_branch and new_has_branch:
            logger.info("既存ファイルに事業所カラムを追加します")
            existing_df.insert(0, '事業所', '')
            # 学校名から事業所を補完
            existing_df['事業所'] = existing_df['学校名'].map(branch_mapping).fillna('')
            existing_has_branch = True

        # 既存ファイルに事業所カラムがあるが空の場合、新しいデータで補完
        if existing_has_branch and new_has_branch:
            mask = (existing_df['事業所'].isna()) | (existing_df['事業所'] == '')
            existing_df.loc[mask, '事業所'] = existing_df.loc[mask, '学校名'].map(branch_mapping).fillna('')

        # カラム設定
        if new_has_branch:
            event_cols = ['事業所', '学校名', 'イベント名', 'イベント開始日', '売り上げ']
            key_cols = ['学校名', 'イベント名', 'イベント開始日']  # マージは学校名ベース
        else:
            event_cols = ['学校名', 'イベント名', 'イベント開始日', '売り上げ']
            key_cols = ['学校名', 'イベント名', 'イベント開始日']

        new_data = new_df[[c for c in event_cols if c in new_df.columns]].copy()
        new_data = new_data.rename(columns={'売り上げ': self.month_col_name})

        if '総計' in existing_df.columns:
            existing_df = existing_df.drop(columns=['総計'])

        if self.month_col_name in existing_df.columns:
            existing_df = existing_df.drop(columns=[self.month_col_name])

        # 既存ファイルのキーカラムを新しいデータに合わせる
        actual_key_cols = [c for c in key_cols if c in existing_df.columns and c in new_data.columns]

        if 'イベント開始日' in existing_df.columns:
            existing_df['イベント開始日'] = pd.to_datetime(existing_df['イベント開始日']).dt.date
        if 'イベント開始日' in new_data.columns:
            new_data['イベント開始日'] = pd.to_datetime(new_data['イベント開始日']).dt.date

        # 事業所カラムがある場合、マージ後に事業所を更新するため一旦除外
        if '事業所' in new_data.columns:
            new_data_for_merge = new_data.drop(columns=['事業所'])
        else:
            new_data_for_merge = new_data

        merged = pd.merge(existing_df, new_data_for_merge, on=actual_key_cols, how='outer')

        # マージ後に事業所が空のレコードを補完
        if '事業所' in merged.columns and branch_mapping:
            mask = (merged['事業所'].isna()) | (merged['事業所'] == '')
            merged.loc[mask, '事業所'] = merged.loc[mask, '学校名'].map(branch_mapping).fillna('')

        merged = self._reorder_month_columns(merged, actual_key_cols, sheet_type='event')

        month_cols = [col for col in merged.columns if '年' in col and '月分' in col]
        merged['総計'] = merged[month_cols].fillna(0).sum(axis=1).astype(int)

        return merged

    def _reorder_month_columns(self, df: pd.DataFrame, key_cols: list, sheet_type: str = 'event') -> pd.DataFrame:
        """月列を年月順に並べ替え、シートタイプに応じてカラム順序を設定

        Args:
            df: 対象のDataFrame
            key_cols: キーカラムのリスト
            sheet_type: 'school'（学校別）または 'event'（イベント別）
        """
        month_cols = [col for col in df.columns if '年' in col and '月分' in col]

        def parse_month_col(col):
            try:
                parts = col.replace('年', ' ').replace('月分', '').split()
                return (int(parts[0]), int(parts[1]))
            except Exception:
                return (9999, 99)

        month_cols_sorted = sorted(month_cols, key=parse_month_col)

        # シートタイプに応じて固定カラムの順序を定義
        if sheet_type == 'school':
            # 学校別: 担当者・写真館・学校名の順
            fixed_col_order = ['担当者', '写真館', '学校名']
        else:
            # イベント別: 事業所・学校名・イベント名・イベント開始日の順
            fixed_col_order = ['事業所', '学校名', 'イベント名', 'イベント開始日']

        other_cols = []
        for col in fixed_col_order:
            if col in df.columns:
                other_cols.append(col)

        # 固定カラムに含まれない他のカラムを追加
        for col in df.columns:
            if col not in other_cols and col not in month_cols and col != '総計':
                other_cols.append(col)

        new_order = other_cols + month_cols_sorted
        if '総計' in df.columns:
            new_order.append('総計')

        return df[new_order]

    def _apply_styles(self):
        """Excelファイルにスタイルを適用"""
        wb = openpyxl.load_workbook(self.output_path)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # ヘッダースタイル
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # データセルの罫線
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # 列幅の自動調整
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                            max_length = max(max_length, cell_length)
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(self.output_path)
